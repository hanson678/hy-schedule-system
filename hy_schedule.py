# -*- coding: utf-8 -*-
"""河源排期写入模块：多文件独立排期，统一列格式(C1-C14 + C25-C28)

关键设计：
- openpyxl只读建全局索引(跨所有(修改)sheet)
- WPS COM写入修改单(多文件分别打开)
- 新单生成到独立Excel供复制粘贴
- 跨文件重复货号由前端弹窗选择后提交
- 列布局复用总排期COL字典(用户已统一列顺序)
"""
import os
import re
import json
import shutil
import logging
from datetime import datetime, timedelta

# 河源排期列号（与(修改)sheet一致，跟总排期完全匹配）
# C1=来单日期 C2=客户名称 C3=走货国 C4=合同号 C5=客PO C6=SKU C7=产品货号 C8=产品名称
# C9=PO数量 C10=内箱 C11=外箱 C12=总箱 C13=CRD C14=验货日期
# C25=备注 C26=跟单 C27=单价HK$ C28=金额HK$
COL = {
    'po_date': 1, 'customer': 2, 'dest': 3, 'po': 4, 'cpo': 5,
    'sku_line': 6, 'item': 7, 'cn_name': 8,
    'qty': 9, 'inner': 10, 'outer': 11, 'total_box': 12,
    'ship_date': 13, 'insp_date': 14,
    'remark': 25, 'from_person': 26, 'price': 27, 'amount': 28,
}

BLUE_COM = 15773696  # RGB(0,176,240)

# 加载AutoFilter补丁（WPS文件兼容）
import openpyxl.descriptors.base as _db
_orig_mp_set = _db.MatchPattern.__set__
def _lenient_mp_set(self, instance, value):
    try:
        _orig_mp_set(self, instance, value)
    except ValueError:
        instance.__dict__[self.name] = None
_db.MatchPattern.__set__ = _lenient_mp_set


def _normalize_po(v):
    """PO号标准化：去掉.0后缀、去空白"""
    s = str(v or '').strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _item_upper(v):
    """货号标准化：去空白，大写"""
    return re.sub(r'[\s\n]+', '', str(v or '')).strip().upper()


# 辅助sheet关键词黑名单（命中任一子串 → 不是排期sheet）
# 注意：'MA' 不能作为普通子串（会误伤 MAIN/MATCH 等），用正则单独处理
AUX_KEYWORDS = (
    '货价',
    '取消',
    '已完成', '已验货', '已发货',
    '数级表',
    '汇总',              # 汇总 / 接单汇总 / 接金汇总
    '尺码',
    '产值走货',
    '产能计划',
    '提前单',
    '发票描述',
    '理论回数',
    '展示架',
    'SUMMARY',
    'STD',
    '旧',
)

_SHEET_DEFAULT_RE = re.compile(r'^SHEET\d*$', re.I)
# MA 用词边界匹配，避免误伤 MAIN/MATCH/GAMA 等合法词
# 匹配条件：MA 前后不能是 ASCII 字母（允许中文/数字/空白/标点包围）
_MA_RE = re.compile(r'(?<![A-Za-z])MA(?![A-Za-z])')
# OLD 同理
_OLD_RE = re.compile(r'(?<![A-Za-z])OLD(?![A-Za-z])', re.I)


def _is_aux_sheet(sheet_name):
    """判断是否为辅助sheet（非排期）"""
    if not sheet_name:
        return True
    sn = str(sheet_name).strip()
    if not sn:
        return True
    # openpyxl 默认命名 Sheet / Sheet1 / Sheet2...
    if _SHEET_DEFAULT_RE.match(sn):
        return True
    up = sn.upper()
    # MA / OLD 边界匹配
    if _MA_RE.search(up) or _OLD_RE.search(up):
        return True
    return any(k.upper() in up for k in AUX_KEYWORDS)


def _pick_target_sheets(sheet_names):
    """从一个文件的 sheet 列表中挑出目标排期 sheet 列表（文件级判定）

    规则：
    1. 优先：含"修改"且非辅助的 sheet → 全部返回（兼容现有命名约定）
    2. 兜底：排除辅助 sheet 后，如果只剩一个候选 → 返回这一个
       （覆盖 92120 这类没有"修改"sheet 的文件）
    3. 否则返回空列表（保守策略，避免误写多候选文件）
    """
    if not sheet_names:
        return []

    # 1. 优先含"修改"且非辅助
    mod_sheets = [s for s in sheet_names if '修改' in s and not _is_aux_sheet(s)]
    if mod_sheets:
        return mod_sheets

    # 2. 非辅助 sheet 只剩一个 → 用它
    non_aux = [s for s in sheet_names if not _is_aux_sheet(s)]
    if len(non_aux) == 1:
        return non_aux

    # 3. 无法判定
    return []


def _is_target_sheet(sheet_name):
    """兼容旧调用：sheet 名级别判断（只看是否含'修改'）

    注意：此函数已弃用于主流程，build_global_index 改用 _pick_target_sheets
    保留仅供外部代码向后兼容调用。
    """
    return '修改' in str(sheet_name or '')


_WEEKDAY_CN = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']


def _fmt_date(dt):
    """日期格式化：2026/5/25周一"""
    if not dt or not hasattr(dt, 'strftime'):
        return ''
    return f"{dt.year}/{dt.month}/{dt.day}{_WEEKDAY_CN[dt.weekday()]}"


def _date_serial(dt):
    """datetime转Excel序列号"""
    if not dt or not hasattr(dt, 'year'):
        return None
    return (dt - datetime(1899, 12, 30)).days


def _is_fuggler(sku):
    """Fuggler系列判断（157开头 或 125160/125169）"""
    s = _item_upper(sku)
    base = re.match(r'(\d+)', s)
    if not base:
        return False
    num = base.group(1)
    return num.startswith('157') or num.startswith('125160') or num.startswith('125169')


def _calc_inspection(ship_dt, sku='', wb_name=''):
    """验货日期计算
    Fuggler系列: 出货-2天; 其他: 出货-4天
    HY(河源): 周六/周日都不能验货 → 周五/周一
    非HY: 只有周日不能验货 → 周一
    Fuggler判断：SKU前缀(157/125160/125169) 或 文件名含'fuggler'
    HY判断：文件名含'hy'或'河源'（大小写不敏感）
    """
    if not ship_dt or not hasattr(ship_dt, 'year'):
        return None
    _fn = str(wb_name or '').lower()
    is_fuggler = _is_fuggler(sku) or 'fuggler' in _fn
    is_hy = any(k in _fn for k in ('hy', '河源'))
    days_before = 2 if is_fuggler else 4
    insp_dt = ship_dt - timedelta(days=days_before)
    if is_hy:
        if insp_dt.weekday() == 5:  # 周六 → 周五
            insp_dt -= timedelta(days=1)
        elif insp_dt.weekday() == 6:  # 周日 → 周一
            insp_dt += timedelta(days=1)
    else:
        if insp_dt.weekday() == 6:  # 周日 → 周一
            insp_dt += timedelta(days=1)
    # 安全网：验货期不能晚于或等于出货期
    if insp_dt >= ship_dt:
        insp_dt = ship_dt - timedelta(days=1)
        if insp_dt.weekday() == 6:
            insp_dt -= timedelta(days=2)
        elif is_hy and insp_dt.weekday() == 5:
            insp_dt -= timedelta(days=1)
    return insp_dt


def _load_item_map(data_dir):
    """加载货号→文件映射表"""
    path = os.path.join(data_dir, 'hy_item_map.json')
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logging.error(f'[河源] 映射表加载失败: {e}')
        return {}


def _find_schedule_files(schedule_dir):
    """扫描目录，返回所有.xlsx文件列表"""
    if not os.path.isdir(schedule_dir):
        return []
    return sorted([
        f for f in os.listdir(schedule_dir)
        if f.endswith('.xlsx') and not f.startswith('~$')
    ])


def build_global_index(schedule_dir):
    """跨所有(修改)sheet构建索引
    Returns: {
        'index': {(po, item_upper): [(file, sheet, row), ...]},
        'cn_names': {item_base: cn_name},
        'sheet_info': {(file, sheet): {'header_row': N, 'max_row': M}},
        'files_scanned': N,
        'sheets_scanned': N,
    }
    """
    import openpyxl

    index = {}
    cn_names = {}
    sheet_info = {}
    files_scanned = 0
    sheets_scanned = 0

    files = _find_schedule_files(schedule_dir)

    for fn in files:
        fpath = os.path.join(schedule_dir, fn)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            logging.warning(f'[河源索引] 打开失败 {fn}: {e}')
            continue

        # 筛选目标sheet（文件级规则：优先"修改"，否则唯一非辅助候选）
        targets = _pick_target_sheets(wb.sheetnames)
        if not targets:
            logging.info(f'[河源索引] 跳过(无可识别排期sheet) {fn}: {wb.sheetnames}')
            wb.close()
            continue

        files_scanned += 1
        for sn in targets:
            try:
                ws = wb[sn]
                # 找表头行（含'产品货号'）
                header_row = None
                for r in range(1, 8):
                    for c in range(1, 20):
                        v = str(ws.cell(r, c).value or '')
                        if '产品货号' in v:
                            header_row = r
                            break
                    if header_row:
                        break
                if not header_row:
                    continue

                sheets_scanned += 1
                max_row = header_row
                empty_count = 0

                # 用enumerate避免EmptyCell的.row属性问题
                for offset, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=30)):
                    r = header_row + 1 + offset
                    po_val = _normalize_po(row[COL['po'] - 1].value if len(row) >= COL['po'] else None)
                    item_val = _item_upper(row[COL['item'] - 1].value if len(row) >= COL['item'] else None)
                    cn_val = str((row[COL['cn_name'] - 1].value if len(row) >= COL['cn_name'] else None) or '').strip()

                    if not item_val:
                        empty_count += 1
                        if empty_count > 50:
                            break
                        continue
                    empty_count = 0
                    max_row = r

                    # 索引：(po, item) → [(file, sheet, row)]
                    if po_val:
                        key = (po_val, item_val)
                        if key not in index:
                            index[key] = []
                        index[key].append((fn, sn, r))

                    # 中文名索引
                    if cn_val and any('\u4e00' <= c <= '\u9fff' for c in cn_val):
                        base = re.match(r'(\d+[A-Za-z]*\d*)', item_val)
                        if base:
                            cn_names[base.group(1).upper()] = cn_val

                sheet_info[(fn, sn)] = {
                    'header_row': header_row,
                    'max_row': max_row,
                }
            except Exception as e:
                logging.warning(f'[河源索引] 处理sheet失败 {fn}/{sn}: {e}')

        wb.close()

    logging.info(f'[河源索引] 完成：{files_scanned}个文件, {sheets_scanned}个sheet, '
                 f'{len(index)}条(PO+货号)记录, {len(cn_names)}个中文名')

    return {
        'index': index,
        'cn_names': cn_names,
        'sheet_info': sheet_info,
        'files_scanned': files_scanned,
        'sheets_scanned': sheets_scanned,
    }


def _prepare_line_data(order, ln, ship_dt_header, full_note, wb_name=''):
    """从PO line组装写入数据
    wb_name: 当前目标排期文件名，用于 _calc_inspection 判断 HY / Fuggler
    """
    sku_spec = ln.get('sku_spec', '') or ln.get('sku', '')
    qty = ln.get('qty', 0) or 0
    price = ln.get('price', 0) or 0
    inner_pcs = ln.get('inner_pcs', 0) or 0
    outer_qty = ln.get('outer_qty', 0) or 0
    customer_po = ln.get('customer_po', '')
    is_pallet = ln.get('is_pallet', False)
    pallet_count = ln.get('pallet_count', 0) or 0
    is_mixed = ln.get('is_mixed_carton', False)
    carton_count = ln.get('carton_count', 0) or 0
    line_no = ln.get('line_no', '')

    # 外箱
    if is_pallet and pallet_count > 0:
        if qty > pallet_count:
            outer_qty = qty // pallet_count
    elif is_mixed and carton_count > 0:
        outer_qty = qty // carton_count

    # 总箱
    if is_pallet and pallet_count > 0:
        total_ctns = pallet_count
    elif is_mixed and carton_count > 0:
        total_ctns = carton_count
    elif outer_qty > 0:
        total_ctns = qty // outer_qty if qty else 0
    else:
        total_ctns = ln.get('total_ctns', 0) or 0

    # 金额
    if (is_pallet and pallet_count > 0) or (is_mixed and carton_count > 0):
        total_hkd = total_ctns * price
    else:
        total_hkd = qty * price

    # 日期
    line_ship = ln.get('delivery', '')
    line_ship_dt = None
    if line_ship:
        try:
            line_ship_dt = datetime.strptime(str(line_ship)[:10], '%Y-%m-%d')
        except Exception:
            line_ship_dt = ship_dt_header
    else:
        line_ship_dt = ship_dt_header
    insp_dt = _calc_inspection(line_ship_dt, sku_spec, wb_name)

    f_sku = f"{_normalize_po(order.get('po_number',''))}-{line_no}" if line_no else ''

    return {
        'sku_spec': sku_spec,
        'qty': qty, 'price': price, 'inner_pcs': inner_pcs,
        'outer_qty': outer_qty, 'total_ctns': total_ctns, 'total_hkd': total_hkd,
        'customer_po': customer_po, 'line_no': line_no, 'f_sku': f_sku,
        'line_ship_dt': line_ship_dt, 'insp_dt': insp_dt,
    }


def _extract_header(order):
    """提取PO header信息"""
    header = order.get('header') or order
    po = _normalize_po(header.get('po_number', '') or order.get('po_number', ''))
    po_date = header.get('po_date', '') or order.get('po_date', '')
    customer = header.get('customer', '') or order.get('customer', '')
    dest = header.get('destination_cn', '') or order.get('destination_cn', '')
    from_person = header.get('from_person', '') or order.get('from_person', '')
    ship_date_str = header.get('ship_date', '') or order.get('ship_date', '')

    tc = header.get('tracking_code', '') or order.get('tracking_code', '') or ''
    pi = header.get('packaging_info', '') or order.get('packaging_info', '') or ''
    rm = header.get('remark', '') or order.get('remark', '') or ''
    note_parts = []
    if tc: note_parts.append(tc)
    if pi: note_parts.append(f'Packaging Info: {pi}')
    if rm: note_parts.append(f'Remark: {rm}')
    full_note = '\n'.join(note_parts)

    ship_dt = None
    if ship_date_str:
        try:
            ship_dt = datetime.strptime(str(ship_date_str)[:10], '%Y-%m-%d')
        except Exception:
            pass

    po_date_dt = None
    if po_date:
        try:
            po_date_dt = datetime.strptime(str(po_date)[:10], '%Y-%m-%d')
        except Exception:
            pass

    return {
        'po': po, 'po_date': po_date, 'po_date_dt': po_date_dt,
        'customer': customer, 'dest': dest, 'from_person': from_person,
        'ship_dt': ship_dt, 'full_note': full_note,
    }


def analyze_orders(schedule_dir, orders):
    """分析PO：分类为修改单/新单，检测跨文件重复货号

    Returns: {
        'modifications': [{order_idx, line_idx, file, sheet, row, item, po}, ...],
        'new_lines': [{order_idx, line_idx, item, po, candidate_files}, ...],
        'ambiguous': [{order_idx, line_idx, item, po, candidates: [...]}, ...],
        'unknown': [{order_idx, line_idx, item, po}, ...],
        'index_info': {files_scanned, sheets_scanned},
    }
    """
    data_dir = os.path.join(os.path.dirname(__file__), 'data')
    item_map = _load_item_map(data_dir)
    idx_data = build_global_index(schedule_dir)
    index = idx_data['index']

    modifications = []
    new_lines = []
    ambiguous = []
    unknown = []

    for oi, order in enumerate(orders):
        hdr = _extract_header(order)
        po = hdr['po']
        lines = order.get('lines', [])
        for li, ln in enumerate(lines):
            sku_spec = ln.get('sku_spec', '') or ln.get('sku', '')
            item_upper = _item_upper(sku_spec)
            if not item_upper:
                continue

            # Step 1: 索引查找 (PO+货号)
            key = (po, item_upper)
            if key in index:
                for fn, sn, r in index[key]:
                    modifications.append({
                        'order_idx': oi, 'line_idx': li,
                        'file': fn, 'sheet': sn, 'row': r,
                        'item': sku_spec, 'po': po,
                    })
                continue

            # Step 2: 货号映射查找（拷贝防止污染item_map）
            candidates = list(item_map.get(item_upper, []))
            if not candidates:
                # 前缀兜底
                base = re.match(r'(\d+[A-Za-z]*\d*)', item_upper)
                if base:
                    for k, v in item_map.items():
                        k_base = re.match(r'(\d+[A-Za-z]*\d*)', k)
                        if k_base and k_base.group(1) == base.group(1):
                            for cand in v:
                                if cand not in candidates:
                                    candidates.append(cand)

            if not candidates:
                unknown.append({
                    'order_idx': oi, 'line_idx': li,
                    'item': sku_spec, 'po': po,
                })
            elif len(set((c['file'], c['sheet']) for c in candidates)) == 1:
                # 唯一定位
                new_lines.append({
                    'order_idx': oi, 'line_idx': li,
                    'item': sku_spec, 'po': po,
                    'file': candidates[0]['file'], 'sheet': candidates[0]['sheet'],
                })
            else:
                # 多文件 → 待用户选择
                ambiguous.append({
                    'order_idx': oi, 'line_idx': li,
                    'item': sku_spec, 'po': po,
                    'candidates': candidates,
                })

    return {
        'modifications': modifications,
        'new_lines': new_lines,
        'ambiguous': ambiguous,
        'unknown': unknown,
        'cn_names': idx_data['cn_names'],  # 供write_orders复用，避免重复扫描
        'index_info': {
            'files_scanned': idx_data['files_scanned'],
            'sheets_scanned': idx_data['sheets_scanned'],
        },
    }


def write_orders(schedule_dir, orders, ambiguous_selections=None, export_dir=None):
    """执行写入：修改单COM写入，新单生成Excel

    ambiguous_selections: {f'{order_idx}_{line_idx}': {'file':..., 'sheet':...}, ...}
                         前端弹窗后提交的用户选择
    Returns: {'ok', 'modified', 'new_count', 'msg', 'export_file',
              'mod_details', 'new_details', 'unknown'}
    """
    if ambiguous_selections is None:
        ambiguous_selections = {}

    # Step 1: 分析
    analysis = analyze_orders(schedule_dir, orders)

    # 合并用户选择的ambiguous到new_lines
    for amb in analysis['ambiguous']:
        sel_key = f"{amb['order_idx']}_{amb['line_idx']}"
        sel = ambiguous_selections.get(sel_key)
        if sel:
            analysis['new_lines'].append({
                'order_idx': amb['order_idx'],
                'line_idx': amb['line_idx'],
                'item': amb['item'], 'po': amb['po'],
                'file': sel['file'], 'sheet': sel['sheet'],
            })

    # Step 2: 按文件分组修改单
    mods_by_file = {}
    for m in analysis['modifications']:
        mods_by_file.setdefault(m['file'], []).append(m)

    # Step 3: 中文名映射（复用analysis的结果，不重复扫描）
    cn_names = dict(analysis.get('cn_names', {}))
    try:
        cn_map_path = os.path.join(os.path.dirname(__file__), 'data', 'item_cn_name_map.json')
        if os.path.exists(cn_map_path):
            with open(cn_map_path, 'r', encoding='utf-8') as f:
                raw = json.load(f)
            for k, v in raw.items():
                if not k.startswith('_') and k.upper() not in cn_names:
                    cn_names[k.upper()] = v.get('cn_name', '') if isinstance(v, dict) else str(v)
    except Exception:
        pass

    # Step 4: 执行COM修改
    mod_count = 0
    mod_details = []

    if mods_by_file:
        import pythoncom
        import win32com.client
        pythoncom.CoInitialize()
        wps = None
        try:
            wps = win32com.client.Dispatch('Ket.Application')
            wps.Visible = False
            wps.DisplayAlerts = False

            for fn, mods in mods_by_file.items():
                fpath = os.path.join(schedule_dir, fn)
                if not os.path.exists(fpath):
                    continue
                # 备份
                try:
                    shutil.copy2(fpath, fpath + '.bak')
                except Exception as e:
                    logging.warning(f'[河源] 备份失败 {fn}: {e}')

                wb = None
                try:
                    wb = wps.Workbooks.Open(fpath)
                    if wb.ReadOnly:
                        logging.warning(f'[河源] {fn}只读，跳过')
                        wb.Close(SaveChanges=False)
                        continue

                    # 按sheet分组
                    mods_by_sheet = {}
                    for m in mods:
                        mods_by_sheet.setdefault(m['sheet'], []).append(m)

                    file_changed = False
                    for sn, sheet_mods in mods_by_sheet.items():
                        ws = None
                        for i in range(1, wb.Sheets.Count + 1):
                            if wb.Sheets(i).Name == sn:
                                ws = wb.Sheets(i)
                                break
                        if not ws:
                            continue

                        for m in sheet_mods:
                            r = m['row']
                            oi, li = m['order_idx'], m['line_idx']
                            order = orders[oi]
                            hdr = _extract_header(order)
                            ln_data = _prepare_line_data(
                                order, order['lines'][li], hdr['ship_dt'], hdr['full_note'],
                                wb_name=fn
                            )

                            # 要写入的字段
                            # 注意：C12总箱 和 C28金额 是排期中的公式（=I/K 和 =AA*I），
                            # 改数量后自动重算，修改单绝对不写这两列以免破坏公式
                            # C27单价 也可能是 VLOOKUP 公式 → 改为"只补不覆盖"（见下方判断）
                            updates = [(COL['qty'], ln_data['qty'])]
                            if ln_data['inner_pcs']:
                                updates.append((COL['inner'], ln_data['inner_pcs']))
                            if ln_data['outer_qty']:
                                updates.append((COL['outer'], ln_data['outer_qty']))
                            if ln_data['customer_po']:
                                updates.append((COL['cpo'], ln_data['customer_po']))
                            if ln_data['line_ship_dt']:
                                updates.append((COL['ship_date'], _date_serial(ln_data['line_ship_dt'])))
                            if ln_data['insp_dt']:
                                updates.append((COL['insp_date'], _date_serial(ln_data['insp_dt'])))
                            if ln_data['price']:
                                updates.append((COL['price'], round(ln_data['price'], 4)))

                            COL_NAMES = {
                                COL['qty']: '数量', COL['inner']: '内箱', COL['outer']: '外箱',
                                COL['cpo']: '客PO',
                                COL['ship_date']: 'CRD', COL['insp_date']: '验货',
                                COL['price']: '单价',
                            }
                            changes = []
                            for col_num, new_val in updates:
                                old_val = ws.Cells(r, col_num).Value
                                # C27单价：只补不覆盖（保护 VLOOKUP 公式和已手填的单价）
                                # 用 float 判定而非 str：VLOOKUP 返回 #N/A 时 old_val 是 pywin32 错误对象，
                                # str() 后非空会被误判为"有值"，导致 N/A 永远无法补价
                                if col_num == COL['price']:
                                    _has_price = False
                                    try:
                                        if old_val is not None and float(old_val) > 0:
                                            _has_price = True
                                    except Exception:
                                        # 兜底 pywin32 访问 CVErr 单元格时可能抛出的 com_error
                                        pass
                                    if _has_price:
                                        continue
                                old_display = old_val
                                if hasattr(old_val, 'year'):
                                    try:
                                        old_display = _fmt_date(old_val.replace(tzinfo=None))
                                        old_val = (old_val.replace(tzinfo=None) - datetime(1899, 12, 30)).days
                                    except Exception:
                                        pass
                                try:
                                    if old_val is not None and new_val is not None:
                                        o_f = float(old_val)
                                        n_f = float(new_val)
                                        if abs(o_f - n_f) < 0.0001:
                                            continue
                                except (ValueError, TypeError):
                                    pass
                                o_s = str(old_val or '').strip()
                                n_s = str(new_val or '').strip()
                                if o_s.endswith('.0'): o_s = o_s[:-2]
                                if n_s.endswith('.0'): n_s = n_s[:-2]
                                if o_s == n_s:
                                    continue
                                ws.Cells(r, col_num).Value = new_val
                                ws.Cells(r, col_num).Interior.Color = BLUE_COM
                                col_label = COL_NAMES.get(col_num, f'列{col_num}')
                                if col_num in (COL['ship_date'], COL['insp_date']):
                                    new_display = _fmt_date(datetime(1899, 12, 30) + timedelta(days=int(new_val))) if new_val else ''
                                    old_show = str(old_display or '')
                                else:
                                    old_show = o_s
                                    new_display = n_s
                                changes.append(f'{col_label} {old_show}→{new_display}')

                            if changes:
                                mod_count += 1
                                file_changed = True
                                mod_details.append({
                                    'item': m['item'], 'row': r, 'po': m['po'],
                                    'file': fn, 'sheet': sn,
                                    'changes': changes,
                                })

                    if file_changed:
                        wb.Save()
                        logging.info(f'[河源] {fn} 已保存')
                    wb.Close(SaveChanges=False)
                except Exception as e:
                    logging.error(f'[河源] 处理{fn}失败: {e}')
                    if wb:
                        try: wb.Close(SaveChanges=False)
                        except: pass
        finally:
            try:
                if wps: wps.Quit()
            except: pass
            pythoncom.CoUninitialize()

    # Step 5: 新单收集 → 生成Excel
    new_rows = []
    new_details = []
    for nl in analysis['new_lines']:
        oi, li = nl['order_idx'], nl['line_idx']
        order = orders[oi]
        hdr = _extract_header(order)
        ln_data = _prepare_line_data(order, order['lines'][li], hdr['ship_dt'], hdr['full_note'],
                                     wb_name=nl['file'])

        item_base = re.match(r'(\d+[A-Za-z]*\d*)', _item_upper(nl['item']))
        cn_name = cn_names.get(item_base.group(1).upper(), '') if item_base else ''

        # 只有当 outer 和 price 有效值时才生成公式占位符，否则留空避免 #DIV/0! / #VALUE!
        _has_outer = bool(ln_data['outer_qty'])
        _has_price = bool(ln_data['price'])
        new_rows.append({
            'target_file': nl['file'], 'target_sheet': nl['sheet'],
            'po_date': hdr['po_date_dt'], 'customer': hdr['customer'], 'dest': hdr['dest'],
            'po': hdr['po'], 'cpo': ln_data['customer_po'], 'sku_line': ln_data['f_sku'],
            'item': ln_data['sku_spec'], 'cn_name': cn_name,
            'qty': ln_data['qty'], 'inner': ln_data['inner_pcs'],
            'outer': ln_data['outer_qty'],
            # 总箱和金额用公式占位符，在 _generate_new_excel 里替换成实际公式
            'total_box': '__FORMULA_TOTAL_BOX__' if _has_outer else '',
            'ship_date': ln_data['line_ship_dt'], 'insp_date': ln_data['insp_dt'],
            'remark': hdr['full_note'],
            'from_person': hdr['from_person'].split('/')[0].strip() if hdr['from_person'] else '',
            'price': round(ln_data['price'], 4) if _has_price else '',
            'amount': '__FORMULA_AMOUNT__' if _has_price else '',
        })
        new_details.append(f"{nl['item']} → {nl['file']}[{nl['sheet']}]")

    # Step 5.5: 收集 unknown(识别不到排期的货号),追加到 new_rows 独立 sheet
    unknown_count = len(analysis['unknown'])
    for uk in analysis['unknown']:
        oi, li = uk['order_idx'], uk['line_idx']
        order = orders[oi]
        hdr = _extract_header(order)
        # wb_name='' 传给 _prepare_line_data,验货期走非HY规则(无目标文件无从判断)
        ln_data = _prepare_line_data(order, order['lines'][li], hdr['ship_dt'], hdr['full_note'],
                                     wb_name='')

        item_base = re.match(r'(\d+[A-Za-z]*\d*)', _item_upper(uk['item']))
        cn_name = cn_names.get(item_base.group(1).upper(), '') if item_base else ''

        _has_outer = bool(ln_data['outer_qty'])
        _has_price = bool(ln_data['price'])
        new_rows.append({
            'target_file': '未识别货号',  # 固定标记,会单独分到一个 sheet
            'target_sheet': '',
            'po_date': hdr['po_date_dt'], 'customer': hdr['customer'], 'dest': hdr['dest'],
            'po': hdr['po'], 'cpo': ln_data['customer_po'], 'sku_line': ln_data['f_sku'],
            'item': ln_data['sku_spec'], 'cn_name': cn_name,
            'qty': ln_data['qty'], 'inner': ln_data['inner_pcs'],
            'outer': ln_data['outer_qty'],
            'total_box': '__FORMULA_TOTAL_BOX__' if _has_outer else '',
            'ship_date': ln_data['line_ship_dt'], 'insp_date': ln_data['insp_dt'],
            'remark': hdr['full_note'],
            'from_person': hdr['from_person'].split('/')[0].strip() if hdr['from_person'] else '',
            'price': round(ln_data['price'], 4) if _has_price else '',
            'amount': '__FORMULA_AMOUNT__' if _has_price else '',
        })

    export_file = ''
    if new_rows and export_dir:
        export_file = _generate_new_excel(new_rows, export_dir)

    # 真正的新单数 = total - unknown(区分统计)
    real_new_count = len(new_rows) - unknown_count

    parts = []
    if mod_count:
        parts.append(f'修改{mod_count}行(已写入)')
    if real_new_count:
        parts.append(f'新增{real_new_count}行(已生成Excel)')
    if unknown_count:
        parts.append(f'未识别{unknown_count}行(已生成Excel-需手动确认)')
    msg = '、'.join(parts) if parts else '无变化'

    return {
        'ok': True, 'modified': mod_count, 'new_count': real_new_count,
        'unknown_count': unknown_count,
        'msg': msg, 'export_file': export_file,
        'mod_details': mod_details, 'new_details': new_details,
        'unknown': analysis['unknown'],
    }


def _gen_sheet_name(file_name, existing):
    """从排期文件名生成简短sheet名
    规则：去年份/ZURU/生产排期通用词,保留货号+产品名,截断31字符,去重
    """
    base = str(file_name)
    # 去扩展名(可能有 .xlsx新.xlsx 这种奇葩后缀)
    base = re.sub(r'\.xlsx?(新)?\.xlsx?$', '', base, flags=re.IGNORECASE)
    base = re.sub(r'\.xlsx?$', '', base, flags=re.IGNORECASE)
    # 去年份前缀
    base = re.sub(r'^2026[年\s]*', '', base)
    # 去 ZURU 标识
    base = re.sub(r'ZURU\s*', '', base, flags=re.IGNORECASE)
    # 去"生产排期/排期表/排期"通用词
    base = re.sub(r'生产排期|排期表|排期', '', base)
    # #Fuggler# → Fuggler
    base = re.sub(r'#Fuggler#', 'Fuggler', base, flags=re.IGNORECASE)
    # (HY) → HY
    base = re.sub(r'[（(]\s*HY\s*[)）]', 'HY', base, flags=re.IGNORECASE)
    # 先去括号数字 (3)(1)(6)(2) 等
    base = re.sub(r'\(\s*\d+\s*\)', '', base)
    # 再去尾部日期 4-4 / 3-2 等
    base = re.sub(r'\s*\d{1,2}-\d{1,2}\s*\.*\s*$', '', base)
    # 去括号残留（去日期后可能暴露新括号）
    base = re.sub(r'\(\s*\d+\s*\)', '', base)
    # Excel sheet 名禁用字符
    base = re.sub(r'[\\/?*\[\]:]', '', base)
    # 折叠多余空白
    base = re.sub(r'\s+', '', base).strip('-_ ')
    if not base:
        base = 'Sheet'
    # 截断到 31 字符
    if len(base) > 31:
        base = base[:31]
    # 去重
    original = base
    i = 2
    while base in existing:
        suffix = f'_{i}'
        base = (original[:31 - len(suffix)]) + suffix
        i += 1
    return base


def _generate_new_excel(new_rows, output_dir):
    """生成新单Excel:按目标文件分组为多个sheet,每sheet列布局严格对应真实排期 C1~C28
    C1-C14 数据列 | C15-C24 空白(跟踪列) | C25-C28 备注/跟单/单价/金额
    用户直接选数据行复制粘贴到排期对应位置,列自动对齐,公式相对引用自动跟随
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import OrderedDict

    # 防御性断言:列布局与排期 COL 字典位置一致,否则公式 =I/K/AA 会错位
    assert COL['qty'] == 9, f"COL['qty'] must be 9, got {COL['qty']}"
    assert COL['outer'] == 11, f"COL['outer'] must be 11, got {COL['outer']}"
    assert COL['price'] == 27, f"COL['price'] must be 27, got {COL['price']}"
    assert COL['total_box'] == 12, f"COL['total_box'] must be 12, got {COL['total_box']}"
    assert COL['amount'] == 28, f"COL['amount'] must be 28, got {COL['amount']}"

    # 28列完整布局(key 对应 COL 字典位置);None 表示该位置留空(跟踪列)
    col_layout = [
        ('po_date',   '接单期'),         # C1
        ('customer',  '第三方客户名称'),  # C2
        ('dest',      '走货国'),         # C3
        ('po',        '合同号'),         # C4
        ('cpo',       '客PO'),           # C5
        ('sku_line',  'SKU'),            # C6
        ('item',      '产品货号'),       # C7
        ('cn_name',   '产品名称'),       # C8
        ('qty',       'PO数量(pcs)'),    # C9
        ('inner',     '内箱装箱数'),     # C10
        ('outer',     '外箱装箱数'),     # C11
        ('total_box', '总箱数'),         # C12
        ('ship_date', 'CRD'),            # C13
        ('insp_date', '验货日期'),       # C14
        (None, ''), (None, ''), (None, ''), (None, ''), (None, ''),  # C15-C19 跟踪列
        (None, ''), (None, ''), (None, ''), (None, ''), (None, ''),  # C20-C24 跟踪列
        ('remark',      '备注'),         # C25
        ('from_person', 'ZURU跟单'),     # C26
        ('price',       '单价HK$'),      # C27
        ('amount',      '金额HK$'),      # C28
    ]

    # 样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    # sheet 标签循环色
    TAB_COLORS = ['FF6B6B', 'FFD93D', '6BCB77', '4D96FF', 'C38BFF', 'FF9F43']

    # 列宽(按列号索引,从1开始)
    col_widths = {
        1: 14, 2: 20, 3: 10, 4: 14, 5: 14, 6: 18, 7: 22, 8: 20,
        9: 12, 10: 10, 11: 10, 12: 10, 13: 14, 14: 14,
        25: 30, 26: 12, 27: 12, 28: 14,
    }
    for c in range(15, 25):
        col_widths.setdefault(c, 6)

    DATE_KEYS = {'po_date', 'ship_date', 'insp_date'}

    # 按目标文件分组,保持首次出现顺序
    rows_by_file = OrderedDict()
    for row in new_rows:
        tf = row.get('target_file', '未知文件')
        rows_by_file.setdefault(tf, []).append(row)

    # 空列表早退,避免 wb.save 崩溃(至少需要1个可见sheet)
    if not rows_by_file:
        return ''

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    existing_sheet_names = set()
    tab_idx = 0
    for target_file, rows in rows_by_file.items():
        is_unknown = (target_file == '未识别货号')
        sheet_name = _gen_sheet_name(target_file, existing_sheet_names)
        existing_sheet_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        # 未识别 sheet 强制纯红色标签以警示;其他 sheet 循环色
        if is_unknown:
            ws.sheet_properties.tabColor = 'FF0000'
        else:
            ws.sheet_properties.tabColor = TAB_COLORS[tab_idx % len(TAB_COLORS)]
            tab_idx += 1

        # 未识别 sheet 在第1行插入红字警告,表头下移到第2行,数据从第3行开始
        header_row_num = 2 if is_unknown else 1
        data_start_row = 3 if is_unknown else 2
        if is_unknown:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=28)
            warn_cell = ws.cell(row=1, column=1,
                                value='注意:以下货号未在直查表找到,请手动确认目标排期后再录入(不要直接粘贴)')
            warn_cell.font = Font(name='宋体', size=12, bold=True, color='FF0000')
            warn_cell.alignment = Alignment(horizontal='center', vertical='center')
            warn_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            ws.row_dimensions[1].height = 28

        # 表头
        for ci, (_, label) in enumerate(col_layout, start=1):
            cell = ws.cell(row=header_row_num, column=ci, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin

        # 数据
        for ri, row_data in enumerate(rows, start=data_start_row):
            for ci, (key, _) in enumerate(col_layout, start=1):
                if key is None:
                    # 跟踪列留空但涂蓝底保持视觉完整
                    cell = ws.cell(row=ri, column=ci, value='')
                    cell.fill = blue_fill
                    cell.border = thin
                    continue
                val = row_data.get(key, '')
                # 公式占位符替换:列位置=真实排期,直接用固定列字母 I/K/AA
                if val == '__FORMULA_TOTAL_BOX__':
                    val = f'=I{ri}/K{ri}'
                elif val == '__FORMULA_AMOUNT__':
                    val = f'=AA{ri}*I{ri}'
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(name='宋体', size=11, color='000000')
                cell.fill = blue_fill
                cell.border = thin
                cell.alignment = Alignment(vertical='center', wrap_text=(key == 'remark'))
                if key in DATE_KEYS and val:
                    cell.number_format = 'yyyy/m/d'

        # 列宽
        for c in range(1, 29):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = col_widths.get(c, 10)
        ws.freeze_panes = f'A{data_start_row}'

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(output_dir, exist_ok=True)
    fname = f'河源新单_{ts}.xlsx'
    out_path = os.path.join(output_dir, fname)
    wb.save(out_path)
    wb.close()
    logging.info(f'[河源新单Excel] {out_path},{len(rows_by_file)}个sheet,{len(new_rows)}行')
    return fname
