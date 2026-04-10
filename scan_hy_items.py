# -*- coding: utf-8 -*-
"""扫描河源排期目录，重建hy_item_map.json
使用：python scan_hy_items.py [排期目录]
"""
import os
import re
import sys
import json
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# AutoFilter补丁
import openpyxl.descriptors.base as _db
_orig = _db.MatchPattern.__set__
def _fix(self, inst, val):
    try: _orig(self, inst, val)
    except ValueError: inst.__dict__[self.name] = None
_db.MatchPattern.__set__ = _fix

import openpyxl

# 复用 hy_schedule.py 的 sheet 筛选逻辑，保持唯一事实源
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from hy_schedule import _pick_target_sheets


def scan(schedule_dir):
    mapping = {}
    files = sorted([f for f in os.listdir(schedule_dir)
                    if f.endswith('.xlsx') and not f.startswith('~$')])
    print(f'扫描 {schedule_dir}，共{len(files)}个文件')

    for fn in files:
        fpath = os.path.join(schedule_dir, fn)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f'[ERR] {fn}: {e}')
            continue

        targets = _pick_target_sheets(wb.sheetnames)
        if not targets:
            print(f'  [SKIP] {fn} 无可识别排期sheet: {wb.sheetnames}')
            wb.close()
            continue

        for sn in targets:
            ws = wb[sn]
            # 找表头行
            header_row = None
            for r in range(1, 8):
                for c in range(1, 20):
                    v = str(ws.cell(r, c).value or '')
                    if '产品货号' in v:
                        header_row = r
                        break
                if header_row:
                    break
            if not header_row:
                continue

            items_count = 0
            empty = 0
            for r in range(header_row + 1, 5000):
                v = ws.cell(r, 7).value  # C7=产品货号
                if not v:
                    empty += 1
                    if empty > 30:
                        break
                    continue
                empty = 0
                s = re.sub(r'[\s\n]+', '', str(v)).strip().upper()
                if not s or not re.match(r'\d', s):
                    continue
                items_count += 1
                if s not in mapping:
                    mapping[s] = []
                entry = {'file': fn, 'sheet': sn}
                if entry not in mapping[s]:
                    mapping[s].append(entry)
            print(f'  [{sn}] {items_count}个货号')
        wb.close()

    return mapping


def main():
    default_dir = r'C:\Users\Administrator\Desktop\河源排期新 - 副本'
    schedule_dir = sys.argv[1] if len(sys.argv) > 1 else default_dir

    if not os.path.isdir(schedule_dir):
        print(f'错误：目录不存在 {schedule_dir}')
        sys.exit(1)

    mapping = scan(schedule_dir)

    # 保存
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
    os.makedirs(data_dir, exist_ok=True)
    out_path = os.path.join(data_dir, 'hy_item_map.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    dup_count = sum(1 for v in mapping.values() if len(v) > 1)
    print(f'\n完成：{len(mapping)}个货号，{dup_count}个跨文件重复')
    print(f'保存到：{out_path}')


if __name__ == '__main__':
    main()
